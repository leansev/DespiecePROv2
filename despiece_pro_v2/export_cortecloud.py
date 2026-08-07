# -*- coding: utf-8 -*-
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = [
    'Cantidad',
    'Largo',
    'Ancho',
    'Función',
    'Tapacanto L1',
    'Tapacanto L2',
    'Tapacanto A1',
    'Tapacanto A2',
    'Material',
    'Complemento',
]

COLOR_HEADER_BG = '2F4F7F'
COLOR_HEADER_FG = 'FFFFFF'
COLOR_ROW_ALT = 'EEF2F7'
COLOR_WHITE = 'FFFFFF'


def solid_fill(color):
    return PatternFill(fill_type='solid', fgColor=color)


def load_payload(json_path):
    with open(json_path, encoding='utf-8-sig') as handle:
        return json.load(handle)


def group_modules(rows):
    modules = []
    current_module = None
    current_pieces = []

    for item in rows:
        row_type = item.get('type')
        if row_type == 'module':
            if current_module is not None:
                modules.append((current_module, current_pieces))
            current_module = item.get('label', '')
            current_pieces = []
        elif row_type == 'piece':
            current_pieces.append(item)

    if current_module is not None:
        modules.append((current_module, current_pieces))

    return modules


def piece_placa(piece):
    placa = piece.get('placa_nombre')
    if placa:
        return placa
    espesor = piece.get('espesor', 0)
    return f'{int(espesor)}mm Blanco'


def clean_module_name(label):
    text = label.rsplit('\u2014', 1)[-1].strip()
    return text


def canto_label(item, canto_value):
    v = int(canto_value or 0)
    if v == 0:
        return ''
    if v == 1:
        color = item.get('canto_rojo_color', '')
        espesor = item.get('canto_rojo_espesor', '')
    else:
        color = item.get('canto_azul_color', '')
        espesor = item.get('canto_azul_espesor', '')
    if not espesor:
        return color  # si no hay espesor cargado, al menos poner el color
    return f'{color} {espesor}mm'.strip()


def write_header_row(sheet, row_index):
    fill = solid_fill(COLOR_HEADER_BG)
    font = Font(bold=True, color=COLOR_HEADER_FG)

    for column_index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')


def write_piece_row(sheet, row_index, item, complemento, use_alt_fill):
    values = [
        item['cantidad'],
        item['largo'],
        item['ancho'],
        item['nombre'],
        canto_label(item, item.get('canto_izq', 0)),
        canto_label(item, item.get('canto_der', 0)),
        canto_label(item, item.get('canto_aba', 0)),
        canto_label(item, item.get('canto_arr', 0)),
        piece_placa(item),
        complemento,
    ]
    fill_color = COLOR_ROW_ALT if use_alt_fill else COLOR_WHITE
    fill = solid_fill(fill_color)

    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=value)
        cell.fill = fill


def adjust_column_widths(sheet):
    for column_index, header in enumerate(HEADERS, start=1):
        width = max(len(header) + 2, 12)
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        sheet.column_dimensions[column_letter].width = width


def write_xlsx(output_path, payload):
    modules = group_modules(payload.get('rows', []))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'CorteCloud'

    write_header_row(sheet, 1)
    row_index = 2
    use_alt_fill = False

    for module_label, pieces in modules:
        complemento = clean_module_name(module_label)
        for piece in pieces:
            write_piece_row(sheet, row_index, piece, complemento, use_alt_fill)
            use_alt_fill = not use_alt_fill
            row_index += 1

    adjust_column_widths(sheet)
    workbook.save(output_path)


def main():
    if len(sys.argv) < 3:
        sys.stderr.write('Uso: export_cortecloud.py salida.xlsx datos.json\n')
        sys.exit(1)

    output_path = sys.argv[1]
    json_path = sys.argv[2]
    payload = load_payload(json_path)
    write_xlsx(output_path, payload)


if __name__ == '__main__':
    main()
